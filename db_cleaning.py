import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from feature_syntactic_checker import word_standardization


# Creating backup
def copy_excel_file(source_path, destination_path):
    try:
        shutil.copy2(source_path, destination_path)
        print(f"File copied from {source_path} to {destination_path}")
        return True
    except FileNotFoundError:
        print(f"Source file not found: {source_path}")
        return False
    except Exception as e:
        print(f"Error copying file: {str(e)}")
        return False


# Cleaning db
def db_cleaning(source_file, destination_file):
    df_copy = pd.read_excel(source_file).copy()

    rows_to_delete = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
    df_after_delete = df_copy.drop(rows_to_delete)
    df_after_delete = df_after_delete.drop(df_after_delete.columns[[0, 3, 4]], axis=1)
    df_after_delete.columns = ['No. de solicitud', 'Denominacion', 'Tipo de signo', 'Fecha de presentacion', 'Estado',
                               'Clases de Niza', 'Titular', 'Vigencia', 'No. Certificado']
    df_after_delete.to_excel(destination_file, index=False)

    workbook = load_workbook(destination_file)
    ss_sheet = workbook['Sheet1']
    ss_sheet.title = 'Original_Data'
    workbook.save(destination_file)


# Filtering db
def db_filtering_by_sign_type_and_status(destination_file, sign_type, status):
    df_to_filter = pd.read_excel(destination_file, sheet_name='Original_Data')
    filtered_data = df_to_filter[df_to_filter['Tipo de signo'] == sign_type]
    filtered_data = filtered_data[filtered_data['Estado'] == status]

    workbook = load_workbook(destination_file)
    new_sheet = workbook.create_sheet(title='Filtered_Data')

    for row in dataframe_to_rows(filtered_data, index=False, header=True):
        new_sheet.append(row)

    workbook.save(destination_file)
    standardization_for_brands(destination_file)


# Brands standardized
def standardization_for_brands(destination_file):
    workbook = load_workbook(destination_file)
    worksheet = workbook['Filtered_Data']

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet[f'B{row}']
        if cell.value is not None:
            cell.value = word_standardization(cell.value)

    workbook.save(destination_file)


# Filtering db by syntactic and phonetic similarity
def db_filtering_by_syntactic_and_phonetic_similarity(destination_file, brand_to_compare):
    df_to_filter = pd.read_excel(destination_file, sheet_name='Filtered_Data')
    filtered_data = df_to_filter[df_to_filter[f'Syntactic overlap percentage with {brand_to_compare}'] > 60]
    filtered_data = filtered_data[filtered_data[f'Phonetic similarity percentage with {brand_to_compare}'] > 60]

    workbook = load_workbook(destination_file)
    new_sheet = workbook.create_sheet(title=f'Result for {brand_to_compare}')

    for row in dataframe_to_rows(filtered_data, index=False, header=True):
        new_sheet.append(row)

    workbook.save(destination_file)
    standardization_for_brands(destination_file)
