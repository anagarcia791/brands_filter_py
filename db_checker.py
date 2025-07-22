from openpyxl import load_workbook

from feature_phonetic_checker import get_total_phonetic_score
from feature_syntactic_checker import word_standardization, is_anagram, overlap_percentage


# Syntactic similarity
def checking_syntactic_similarity(destination_file, brand_to_compare):
    standardized_brand = word_standardization(brand_to_compare)

    workbook = load_workbook(destination_file)
    worksheet = workbook['Filtered_Data']

    new_column_index = worksheet.max_column
    worksheet.cell(row=1, column=new_column_index, value=f"Is anagram with {standardized_brand}")

    for row in range(2, worksheet.max_row + 1):
        result = is_anagram(standardized_brand, worksheet.cell(row=row, column=2).value)
        worksheet.cell(row=row, column=new_column_index, value=result)

    worksheet.cell(row=1, column=new_column_index + 1, value=f"Syntactic overlap percentage with {standardized_brand}")

    for row in range(2, worksheet.max_row + 1):
        result = overlap_percentage(standardized_brand, worksheet.cell(row=row, column=2).value)
        worksheet.cell(row=row, column=new_column_index + 1, value=result)

    workbook.save(destination_file)


# Phonetic similarity
def checking_phonetic_similarity(destination_file, brand_to_compare):
    standardized_brand = word_standardization(brand_to_compare)

    workbook = load_workbook(destination_file)
    worksheet = workbook['Filtered_Data']

    new_column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=new_column_index, value=f"Phonetic similarity percentage with {standardized_brand}")

    for row in range(2, worksheet.max_row + 1):
        result = get_total_phonetic_score(standardized_brand, worksheet.cell(row=row, column=2).value)
        worksheet.cell(row=row, column=new_column_index, value=result)

    workbook.save(destination_file)


# Mean value
def mean_value_calculation(destination_file):
    workbook = load_workbook(destination_file)
    worksheet = workbook['Filtered_Data']

    new_column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=new_column_index, value=f"Mean value calculation of syntactic and fonetic result")

    for row in range(2, worksheet.max_row + 1):
        result = (worksheet.cell(row=row, column=10).value + worksheet.cell(row=row, column=11).value) / 2
        worksheet.cell(row=row, column=new_column_index, value=result)

    workbook.save(destination_file)


# Risk perception
def risk_perception(destination_file):
    workbook = load_workbook(destination_file)
    worksheet = workbook['Filtered_Data']

    new_column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=new_column_index, value=f"Risk perception")

    for row in range(2, worksheet.max_row + 1):
        result = ""
        row_value = worksheet.cell(row=row, column=12).value
        if row_value > 79:
            result = "HIGH"
        elif 55 < row_value < 80:
            result = "MEDIUM"
        else:
            result = "LOW"
        worksheet.cell(row=row, column=new_column_index, value=result)

    workbook.save(destination_file)